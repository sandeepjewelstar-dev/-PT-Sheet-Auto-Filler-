"""
Excel Writer Module - Fill Excel PT sheet with extracted data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from pathlib import Path
from config import OUTPUT_DIR, OUTPUT_FILE_PREFIX

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def fill_pt_sheet(excel_template_path, extracted_data):
    """
    Fill Excel PT sheet with extracted data
    
    Args:
        excel_template_path: Path to Excel template file
        extracted_data: Dictionary with extracted data from image
        
    Returns:
        str: Path to filled Excel file
    """
    try:
        logger.info(f"📂 Loading Excel template: {excel_template_path}")
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_template_path)
        ws = wb.active
        
        logger.info("✍️  Filling Excel sheet with extracted data...")
        
        # Fill basic information (adjust cell references based on your template)
        fill_header_info(ws, extracted_data)
        fill_materials(ws, extracted_data)
        fill_stones(ws, extracted_data)
        fill_totals(ws, extracted_data)
        fill_findings_remarks(ws, extracted_data)
        
        # Generate output filename
        product_id = extracted_data.get('product_id', 'OUTPUT')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"{OUTPUT_FILE_PREFIX}{product_id}_{timestamp}.xlsx"
        output_path = OUTPUT_DIR / output_filename
        
        # Save workbook
        logger.info(f"💾 Saving filled Excel sheet: {output_path}")
        wb.save(output_path)
        
        logger.info(f"✅ Excel sheet filled successfully!")
        logger.info(f"   Output file: {output_path}")
        
        return str(output_path)
        
    except Exception as e:
        logger.error(f"❌ Error filling Excel sheet: {str(e)}")
        return None


def fill_header_info(ws, data):
    """Fill header information section"""
    
    # Define cell references (adjust based on your template structure)
    # Example structure - modify according to your Excel template
    
    header_mapping = {
        'B1': 'surface_area',      # Surface Area value
        'D1': 'volume',            # Volume value
        'B2': 'cad_person',        # CAD Person value
        'D2': 'cad_process',       # CAD Process value
        'B3': 'product_id',        # Product ID value
    }
    
    for cell_ref, data_key in header_mapping.items():
        value = data.get(data_key, '')
        if value:
            ws[cell_ref] = value
            logger.debug(f"   {cell_ref} = {value}")


def fill_materials(ws, data):
    """Fill materials section"""
    
    materials = data.get('materials', {})
    
    if not materials:
        logger.warning("⚠️  No materials data to fill")
        return
    
    # Define starting row for materials (adjust based on your template)
    start_row = 8  # Adjust this to match your template
    
    material_labels = [
        'PLATINUM',
        'GWT 18 KT',
        'GWT 14 KT',
        'GWT 10 KT',
        'SILVER'
    ]
    
    for idx, label in enumerate(material_labels):
        row = start_row + idx
        ws[f'A{row}'] = label
        
        # Fill weight if available
        if label in materials:
            weight = materials[label]
            ws[f'B{row}'] = weight
            logger.debug(f"   Row {row}: {label} = {weight} gms")


def fill_stones(ws, data):
    """Fill stone settings section"""
    
    stones = data.get('stones', [])
    
    if not stones:
        logger.warning("⚠️  No stone data to fill")
        return
    
    # Define starting row for stones (adjust based on your template)
    start_row = 15  # Adjust this to match your template
    
    # Column mapping for stones
    stone_columns = {
        'A': 'setting',
        'B': 'shape',
        'C': 'quality',
        'D': 'cut',
        'E': 'mm',
        'F': 'qty',
        'G': 'pts',
        'H': 'weight',
    }
    
    for stone_idx, stone in enumerate(stones):
        row = start_row + stone_idx
        
        for col_letter, data_key in stone_columns.items():
            value = stone.get(data_key, '')
            if value:
                ws[f'{col_letter}{row}'] = value
                logger.debug(f"   {col_letter}{row} = {value}")


def fill_totals(ws, data):
    """Fill total quantities and weights"""
    
    # Find the Total row (usually after all stones)
    # Adjust based on your template
    total_row = 21  # Adjust this to match your template
    
    total_qty = data.get('total_qty', '')
    total_weight = data.get('total_weight', '')
    
    if total_qty:
        ws[f'F{total_row}'] = total_qty
        logger.debug(f"   Total QTY: {total_qty}")
    
    if total_weight:
        ws[f'H{total_row}'] = total_weight
        logger.debug(f"   Total WEIGHT: {total_weight}")


def fill_findings_remarks(ws, data):
    """Fill findings and remarks section"""
    
    # Adjust cell references based on your template
    finding_row = 24  # Adjust this to match your template
    remarks_row = 25  # Adjust this to match your template
    
    finding = data.get('finding', '')
    remarks = data.get('remarks', '')
    
    if finding:
        ws[f'B{finding_row}'] = finding
        logger.debug(f"   Finding: {finding}")
    
    if remarks:
        ws[f'B{remarks_row}'] = remarks
        logger.debug(f"   Remarks: {remarks}")


def validate_filled_sheet(excel_file_path):
    """
    Validate the filled Excel sheet for completeness
    
    Args:
        excel_file_path: Path to filled Excel file
        
    Returns:
        bool: True if valid, False otherwise
    """
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        # Check for required fields
        required_cells = ['B1', 'B2', 'B3']  # Adjust based on your template
        
        missing_fields = []
        for cell in required_cells:
            if ws[cell].value is None:
                missing_fields.append(cell)
        
        if missing_fields:
            logger.warning(f"⚠️  Missing fields: {missing_fields}")
            return False
        
        logger.info("✅ Excel sheet validation passed")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error validating Excel sheet: {e}")
        return False


if __name__ == "__main__":
    # Test filling
    test_template = "templates/PT_Sheet_Template.xlsx"
    test_data = {
        'product_id': 'PLDR5552-PE150',
        'cad_person': 'MAYURI',
        'cad_process': 'Direct Resin (CPX)',
        'surface_area': '856.32',
        'volume': '317',
        'materials': {
            'PLATINUM': '5.87',
            'GWT 18 KT': '4.32',
            'GWT 14 KT': '3.56',
            'GWT 10 KT': '3.15',
            'SILVER': '2.81',
        },
        'stones': [
            {
                'setting': 'Claw prong',
                'shape': 'Pear',
                'quality': 'LGD',
                'cut': '',
                'mm': '9.98*6.44*4.05',
                'qty': '1',
                'pts': '1.51',
                'weight': '1.510',
            }
        ],
        'total_qty': '11',
        'total_weight': '1.822',
        'finding': 'Medium double notch post and back',
        'remarks': 'Sample remarks',
    }
    
    result = fill_pt_sheet(test_template, test_data)
    print(f"\nResult: {result}")
